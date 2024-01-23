from openpyxl import Workbook,load_workbook
class Excel:
    def __init__(self,fayl_nomi=None,sheet=None):
        if fayl_nomi==None:
          self.fayl=Workbook()
          self.jadval=self.fayl.active
        else:
          self.fayl=load_workbook(f"{fayl_nomi}")
          self.jadval=self.fayl[f"{sheet}"]
    def get(self,address):
        return self.jadval[address].value
     
    def update(self,address,qiymat):
        self.jadval[address]=qiymat
        
obj=Excel()        